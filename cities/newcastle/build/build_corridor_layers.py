#!/usr/bin/env python
"""Build the corridor-specific supply layers: A4 (light rail vehicle and dwell),
A2 (signal control, the SCATS proxy) and the A3 GTFS extension tables.

The two highest-leverage unknowns named in proposal section 6.2 -
`dwell_charging_s` and `tsp_enabled` / SCATS phasing - are not published. They
are therefore constructed here as explicit, separately-toggleable additive terms
with sweep ranges, never folded into a single run-time number. Every assumed
value carries source='assumed' and a sweep range, and is mirrored in DECISIONS.md.

Run-time decomposition per segment:
    scheduled = kinematic + boarding_dwell + charging_dwell + signal_delay + recovery
The kinematic term is computed from vehicle dynamics and the true alignment
length, so the residual that must be explained by dwell and signals is derived,
not guessed at wholesale.
"""

# This builder encodes THIS CITY's intervention, corridor or history, so it lives
# with the city rather than in the framework. It still uses the framework's
# generic machinery, which is two directories up.
import os as _os
import sys as _sys
_REPO = _os.path.dirname(_os.path.dirname(_os.path.dirname(
    _os.path.dirname(_os.path.abspath(__file__)))))
_sys.path.insert(0, _os.path.join(_REPO, 'src'))
_sys.path.insert(0, _os.path.join(_REPO, 'src', 'build'))
import city as _city  # noqa: E402
import os
import csv
import json
import math
import zipfile
import io
import collections

import sys as _sys
import registry as _registry  # noqa: E402
CFG = _registry.load()

OUT = _city.path('data/processed/corridor')
os.makedirs(OUT, exist_ok=True)
LRZIP = _city.path('schedules/raw/base2026/lightrail.zip')

# The observed statewide signal inventory. The corridor intersections in the A2
# layer are clusters of OSM traffic-signal nodes and carry no identifier of
# their own, so `scats_site_id` has always been declared and left empty. This is
# the only published source of both the SCATS site number and the date each
# signal was installed (DECISIONS.md 9.24).
SCATS_XLSX = _city.path('data/raw/signals/tfnsw_traffic_lights_location.xlsx')
SCATS_MATCH_M = CFG.get('A.signals.scats_match_radius_m')


def load_scats_inventory():
    """Every signal in the TfNSW inventory, as (lat, lon, site_id, installed).

    Read statewide and matched by distance rather than pre-filtered to a bounding
    box: 4,582 signals against 14 intersections is trivial to scan, and a bbox
    would be one more undeclared constant deciding which observations are
    eligible. Rows without usable coordinates are skipped rather than guessed at.
    """
    import openpyxl
    wb = openpyxl.load_workbook(SCATS_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    hdr = [str(c).strip() if c is not None else '' for c in next(it)]
    ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in it:
        try:
            la, lo = float(r[ix['Latitude']]), float(r[ix['Longitude']])
        except (TypeError, ValueError, KeyError):
            continue
        d = r[ix['Date_Install']]
        out.append((la, lo, str(r[ix['Equipment_ID']]), str(d)[:10] if d else ''))
    wb.close()
    return out


def match_scats(ll, inv):
    """Nearest observed signal to a corridor intersection, or blanks if none.

    An unmatched intersection returns empty fields and is written with
    scats_source='unmatched' rather than being dropped or given a neighbour's
    id: a missing identity must stay visible in the artefact.
    """
    best = None
    for la, lo, site, installed in inv:
        d = hav(ll, (la, lo))
        if best is None or d < best[0]:
            best = (d, site, installed)
    if best is None or best[0] > SCATS_MATCH_M:
        return '', '', ''
    return best[1], round(best[0], 1), best[2]

# --------------------------------------------------------------------------
# A4 vehicle specification - CAF Urbos 100, Newcastle fleet 2151-2156
# Published: 32.966 m, 5 modules, 45 t, 750 V DC, ACR supercapacitor charging
# at stops, wire-free on line. Remaining fields are class-typical for a 33 m
# 100% low-floor 5-module Urbos and are flagged as assumed.
# --------------------------------------------------------------------------
VEHICLE = dict(
    vehicle_type_id='CAF_URBOS_100_NLR',
    model='CAF Urbos 100 (5-module, 100% low floor)',
    fleet_size=6, fleet_numbers='2151-2156',
    length_m=32.966, width_m=2.65, mass_tonnes=45,
    capacity_seated=60, capacity_standing=210, capacity_crush=270,
    capacity_seated_source='assumed', capacity_crush_source='published',
    max_accel_ms2=1.2, max_decel_ms2=1.3, emergency_decel_ms2=2.8,
    max_speed_kmh=70, line_speed_kmh=40,
    door_count_per_side=4, door_width_mm=1300,
    boarding_rate_pax_s=0.6, alighting_rate_pax_s=0.8,
    traction_voltage_v=750, energy_storage='supercapacitor (ACR)',
    charging_mode='pantograph raise to overhead ACR charger at each stop',
    accel_source='assumed', door_source='assumed',
    boarding_rate_source='assumed',
    notes='Wire-free between stops; 750 V DC overhead in depot only.')

# --------------------------------------------------------------------------
# A4 dwell model. dwell_charging_s is the single most consequential assumed
# number in the model (proposal 6.2) and is therefore a separate additive term.
# --------------------------------------------------------------------------
DWELL_DEFAULTS = dict(
    dwell_fixed_s=8.0,              # door open/close + driver reaction
    dwell_fixed_sweep=(5.0, 12.0),
    dwell_charging_s=20.0,          # proposal's own working estimate
    dwell_charging_sweep=(10.0, 35.0),
    dwell_sd_s=6.0,
    distribution_type='lognormal')

# Terminus stops also charge, and hold for layover; intermediate stops are the
# ones that matter for run time.
LR_STOPS_ORDER = ['Newcastle Interchange', 'Honeysuckle', 'Civic',
                  'Crown Street', 'Queens Wharf', 'Newcastle Beach']


def hav(a, b):
    R = 6371000.0
    p1, p2 = math.radians(a[0]), math.radians(b[0])
    dl = math.radians(b[1] - a[1])
    dp = p2 - p1
    return 2 * R * math.asin(math.sqrt(
        math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2))


def kinematic_time(d, v_kmh, a, b):
    """Seconds to cover d metres from rest to rest, trapezoidal or triangular."""
    v = v_kmh / 3.6
    da, db = v * v / (2 * a), v * v / (2 * b)
    if d >= da + db:
        return v / a + v / b + (d - da - db) / v
    vp = math.sqrt(2 * d * a * b / (a + b))
    return vp / a + vp / b


def read_lr():
    z = zipfile.ZipFile(LRZIP)

    def rd(n):
        with z.open(n) as f:
            return list(csv.DictReader(io.TextIOWrapper(f, 'utf-8-sig')))
    return rd('stops.txt'), rd('stop_times.txt'), rd('trips.txt'), rd('shapes.txt')


def sec(t):
    h, m, s = map(int, t.split(':'))
    return h * 3600 + m * 60 + s


def build():
    stops, st, trips, shapes = read_lr()
    sname = {s['stop_id']: s['stop_name'].replace(' Light Rail', '') for s in stops}
    sll = {s['stop_id']: (float(s['stop_lat']), float(s['stop_lon'])) for s in stops}
    tmap = {t['trip_id']: t for t in trips}

    # true alignment length per direction
    byshape = collections.defaultdict(list)
    for r in shapes:
        byshape[r['shape_id']].append(r)
    shp_len = {}
    for sid, rows in byshape.items():
        rows.sort(key=lambda r: int(r['shape_pt_sequence']))
        pts = [(float(r['shape_pt_lat']), float(r['shape_pt_lon'])) for r in rows]
        shp_len[sid] = sum(hav(a, b) for a, b in zip(pts, pts[1:]))

    bytrip = collections.defaultdict(list)
    for r in st:
        bytrip[r['trip_id']].append(r)
    for k in bytrip:
        bytrip[k].sort(key=lambda r: int(r['stop_sequence']))

    # one representative trip per direction
    seen = {}
    for tid, rows in bytrip.items():
        d = tmap[tid]['direction_id']
        if d not in seen:
            seen[d] = (tid, rows)

    seg_rows = []
    for d, (tid, rows) in sorted(seen.items()):
        shid = tmap[tid].get('shape_id')
        # scale straight-line spacing up to the true shape length
        raw = [hav(sll[a['stop_id']], sll[b['stop_id']]) for a, b in zip(rows, rows[1:])]
        scale = (shp_len.get(shid, sum(raw)) / sum(raw)) if sum(raw) else 1.0
        for (a, b), dist in zip(zip(rows, rows[1:]), raw):
            dist = dist * scale
            sched = sec(b['arrival_time']) - sec(a['departure_time'])
            kin = kinematic_time(dist, VEHICLE['line_speed_kmh'],
                                 VEHICLE['max_accel_ms2'], VEHICLE['max_decel_ms2'])
            seg_rows.append(dict(
                direction_id=d, from_stop=sname[a['stop_id']], to_stop=sname[b['stop_id']],
                distance_m=round(dist, 1),
                scheduled_runtime_s=sched,
                kinematic_runtime_s=round(kin, 1),
                residual_s=round(sched - kin, 1),
                implied_mean_speed_kmh=round(dist / sched * 3.6, 1),
                line_speed_kmh=VEHICLE['line_speed_kmh'],
                distance_source='gtfs_shape_scaled',
                kinematic_source='computed'))

    # ---- residual allocation across dwell / signals / recovery ----
    # For each direction: total residual = sum over segments. Intermediate stops
    # carry fixed + charging dwell; the rest is attributed to signal delay and
    # schedule recovery, split by the number of signalised crossings per segment.
    n_int = len(LR_STOPS_ORDER) - 2
    per_dir = collections.defaultdict(float)
    for r in seg_rows:
        per_dir[r['direction_id']] += r['residual_s']
    dwell_per_stop = DWELL_DEFAULTS['dwell_fixed_s'] + DWELL_DEFAULTS['dwell_charging_s']
    alloc = {}
    for d, resid in per_dir.items():
        dwell_total = dwell_per_stop * n_int
        remainder = resid - dwell_total
        alloc[d] = dict(residual_total_s=round(resid, 1),
                        dwell_total_s=round(dwell_total, 1),
                        signal_and_recovery_s=round(remainder, 1),
                        n_intermediate_stops=n_int)

    # ---- A4 stop dwell model ----
    dwell_rows = []
    for i, s in enumerate(LR_STOPS_ORDER):
        terminus = i in (0, len(LR_STOPS_ORDER) - 1)
        dwell_rows.append(dict(
            stop_name=s, stop_seq=i + 1, is_terminus=int(terminus),
            dwell_fixed_s=DWELL_DEFAULTS['dwell_fixed_s'],
            dwell_fixed_sweep_low=DWELL_DEFAULTS['dwell_fixed_sweep'][0],
            dwell_fixed_sweep_high=DWELL_DEFAULTS['dwell_fixed_sweep'][1],
            dwell_boarding_fn='max(pax_board/%.2f, pax_alight/%.2f)' % (
                VEHICLE['boarding_rate_pax_s'], VEHICLE['alighting_rate_pax_s']),
            dwell_charging_s=DWELL_DEFAULTS['dwell_charging_s'],
            dwell_charging_sweep_low=DWELL_DEFAULTS['dwell_charging_sweep'][0],
            dwell_charging_sweep_high=DWELL_DEFAULTS['dwell_charging_sweep'][1],
            dwell_sd_s=DWELL_DEFAULTS['dwell_sd_s'],
            distribution_type=DWELL_DEFAULTS['distribution_type'],
            layover_s=180 if terminus else 0,
            source='assumed',
            acquisition_route='field measurement, or GTFS-Realtime dwell inference'))

    # ---- A2 corridor signal control (SCATS proxy) ----
    # Signalised intersections along the alignment, taken from the OSM signal
    # inventory filtered to a 60 m buffer of the light rail shape.
    sigs = []
    sig_path = _city.path('data/processed/network/A2_signal_nodes_osm.csv')
    if os.path.exists(sig_path):
        pts = []
        for sid, rows in byshape.items():
            rows.sort(key=lambda r: int(r['shape_pt_sequence']))
            pts += [(float(r['shape_pt_lat']), float(r['shape_pt_lon'])) for r in rows]
        for r in csv.DictReader(open(sig_path, encoding='utf-8')):
            ll = (float(r['lat']), float(r['lon']))
            dmin = min(hav(ll, p) for p in pts)
            if dmin <= 60:
                sigs.append((dmin, r, ll))
    # OSM places one signal node per approach; cluster nodes within 45 m into a
    # single intersection so the corridor carries intersections, not stop lines.
    sigs.sort(key=lambda x: x[2][1])   # west to east
    clusters = []
    for dmin, r, ll in sigs:
        for c in clusters:
            if hav(ll, c['centre']) <= 45:
                c['members'].append((dmin, r, ll))
                n = len(c['members'])
                c['centre'] = (sum(m[2][0] for m in c['members']) / n,
                               sum(m[2][1] for m in c['members']) / n)
                break
        else:
            clusters.append({'centre': ll, 'members': [(dmin, r, ll)]})
    sigs = [(min(m[0] for m in c['members']),
             {'node_id': ';'.join(m[1]['node_id'] for m in c['members']),
              'ped_phase_flag': '1' if any(m[1].get('ped_phase_flag') == '1'
                                           for m in c['members']) else '-1',
              'n_approach_nodes': len(c['members'])},
             c['centre']) for c in clusters]
    sigs.sort(key=lambda x: x[2][1])
    scats = load_scats_inventory()
    sig_rows = []
    for i, (dmin, r, ll) in enumerate(sigs):
        site, site_d, installed = match_scats(ll, scats)
        sig_rows.append(dict(
            intersection_id='NLR_SIG_%02d' % (i + 1),
            osm_node_id=r['node_id'], scats_site_id=site,
            scats_match_dist_m=site_d, signal_installed=installed,
            scats_source='observed' if site else 'unmatched',
            n_approach_nodes=r.get('n_approach_nodes',''),
            lat=ll[0], lon=ll[1], dist_to_alignment_m=round(dmin, 1),
            control_type='adaptive',            # SCATS is adaptive by definition
            cycle_time_s=110, cycle_time_sweep_low=80, cycle_time_sweep_high=140,
            n_phases=4, phase_split_pct='45|15|30|10', offset_s=0,
            coordination_group='HUNTER_SCOTT',
            pedestrian_phase_flag=1 if r.get('ped_phase_flag') == '1' else 0,
            ped_clearance_s=8,
            tsp_enabled=0, tsp_type='', tsp_detection_distance_m=0,
            tsp_max_extension_s=0,
            mean_delay_to_tram_s=round(110 * 0.5 * 0.45, 1),
            source='assumed', scenario_variant_ref='S2_base'))
    # S2b variant: full transit signal priority on the corridor
    tsp_rows = []
    for r in sig_rows:
        q = dict(r)
        q.update(tsp_enabled=1, tsp_type='green_extension+early_start',
                 tsp_detection_distance_m=120, tsp_max_extension_s=12,
                 mean_delay_to_tram_s=round(float(r['mean_delay_to_tram_s']) * 0.25, 1),
                 scenario_variant_ref='S2b_full_tsp')
        tsp_rows.append(q)

    # E1 names four signal variants; only two were built. S0_no_tram and
    # S2c_reserved_alignment are the other two, and without them the scenarios
    # that reference them have no signal layer at all.
    #
    # S0_no_tram - the corridor with no tram in it. E1_road_variants sets the
    # cycle to 100 s for the full-capacity network, and there is no tram to delay.
    # S2c_reserved_alignment - the tram is off the street on the former railway
    # alignment, so it meets 60% fewer at-grade signal conflicts (DECISIONS 10).
    # Both are assumed and swept on the same 80-140 s cycle range as S2.
    no_tram_rows = []
    for r in sig_rows:
        q = dict(r)
        q.update(cycle_time_s=100, mean_delay_to_tram_s=0.0,
                 tsp_enabled=0, tsp_type='', source='assumed',
                 scenario_variant_ref='S0_no_tram')
        no_tram_rows.append(q)
    reserved_rows = []
    for r in sig_rows:
        q = dict(r)
        q.update(mean_delay_to_tram_s=round(float(r['mean_delay_to_tram_s']) * 0.40, 1),
                 tsp_enabled=0, tsp_type='', source='assumed',
                 scenario_variant_ref='S2c_reserved_alignment')
        reserved_rows.append(q)
    # S3_brt_priority - the BRT alternative runs in the same lane take as the tram
    # and is given the same priority mechanism, so that S2b and S3 differ in vehicle
    # and dwell rather than in how generously each is signalled. Assumed.
    brt_rows = []
    for r in sig_rows:
        q = dict(r)
        q.update(tsp_enabled=1, tsp_type='green_extension+early_start',
                 tsp_detection_distance_m=120, tsp_max_extension_s=12,
                 mean_delay_to_tram_s=round(float(r['mean_delay_to_tram_s']) * 0.25, 1),
                 source='assumed', scenario_variant_ref='S3_brt_priority')
        brt_rows.append(q)

    def w(name, rows):
        if not rows:
            print('  (empty) %s' % name)
            return
        cols = list(dict.fromkeys(k for x in rows for k in x))
        with open(os.path.join(OUT, name), 'w', newline='', encoding='utf-8') as fh:
            wr = csv.DictWriter(fh, fieldnames=cols, extrasaction='ignore')
            wr.writeheader()
            wr.writerows(rows)
        print('  wrote %-38s %d rows' % (name, len(rows)))

    w('A4_vehicle_spec.csv', [VEHICLE])
    w('A4_stop_dwell_model.csv', dwell_rows)
    w('A4_segment_runtime_decomposition.csv', seg_rows)
    w('A2_signal_control_corridor.csv',
      sig_rows + tsp_rows + no_tram_rows + reserved_rows + brt_rows)

    rep = dict(alignment_length_m={k: round(v, 1) for k, v in shp_len.items()},
               scheduled_end_to_end_s={d: sum(r['scheduled_runtime_s'] for r in seg_rows
                                              if r['direction_id'] == d) for d in per_dir},
               kinematic_end_to_end_s={d: round(sum(r['kinematic_runtime_s'] for r in seg_rows
                                                    if r['direction_id'] == d), 1) for d in per_dir},
               residual_allocation=alloc,
               n_corridor_signals=len(sig_rows),
               # the observed flag reaches the layer (#120): both branches of
               # the ternary above were 1 until 3 Sep 2026
               corridor_intersections_with_pedestrian_phase=sum(
                   1 for r in sig_rows if r.get('scenario_variant_ref') == 'S2_base'
                   and r.get('pedestrian_phase_flag') == 1),
               signal_variants=['S2_base', 'S2b_full_tsp', 'S0_no_tram',
                                'S2c_reserved_alignment', 'S3_brt_priority'],
               charging_dwell_share_of_runtime={
                   d: round(DWELL_DEFAULTS['dwell_charging_s'] * n_int /
                            sum(r['scheduled_runtime_s'] for r in seg_rows
                                if r['direction_id'] == d) * 100, 1) for d in per_dir})
    json.dump(rep, open(os.path.join(OUT, '_corridor_report.json'), 'w', newline='\n'), indent=2)
    print(json.dumps(rep, indent=2))


if __name__ == '__main__':
    build()
